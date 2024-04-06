import time
from datetime import datetime
from pathlib import Path

import openpyxl
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait

time_now = datetime.now()
current_time = time_now.strftime("%H:%M:%S")
print('زمان شروع' + str(current_time))
currentPath = str(Path.cwd())
currentPath = currentPath.replace('\\', '/')
filePath = currentPath + '/' + "addOtherCost.xlsx"
wb = openpyxl.load_workbook(filePath)
sheet_to_focus = 'addOtherCost'
for s in range(len(wb.sheetnames)):
    if wb.sheetnames[s] == sheet_to_focus:
        break
wb.active = s

driver = webdriver.Chrome()
driver.get("https://test.rasatpa.ir/sso/login")
driver.maximize_window()
app = driver.find_element(By.ID, 'app-select')
username = driver.find_element(By.ID, 'username')
password = driver.find_element(By.ID, 'password')
login = driver.find_element(By.ID, 'lbtn')
time.sleep(1)
appName = "کارتابل"
timeOut = 20
app.send_keys(appName)
time.sleep(1)
username.send_keys("1270138758")
# 1288239904
password.send_keys("Ras@4321")
# 1288239904@r1883
time.sleep(1)
login.click()
time.sleep(3)
action = webdriver.ActionChains(driver)
try:
    organizationSelect = driver.find_element(By.XPATH,
                                             "/html/body/div[1]/div[2]/div/div/div/div/div/div/div[2]/div/a/p")
    # / html / body / div[1] / div[2] / div / div / div / div / div / div / div[2] / div / a / p
    organizationSelect.click()
except NoSuchElementException:
    pass
time.sleep(5)
try:
    collapseBtnTaskSearchPanelId = driver.find_element(By.ID, "collapse-btn-taskSearchPanelId")
    collapseBtnTaskSearchPanelId.click()
except NoSuchElementException:
    pass
time.sleep(10)
try:
    barcode = driver.find_element(By.XPATH,
                                  "/html/body/main/div/div/div/div[3]/div/div[2]/div/div[1]/div/div[2]/div["
                                  "2]/div/div/div/form/div/div[2]/div/input")
    barcode.send_keys("1402/38/34669/031/11573")
    # 1402 / 20 / 530 / 59
except NoSuchElementException:
    pass
time.sleep(5)
time.sleep(5)
try:
    taskListGridSearchBtn = driver.find_element(By.ID, 'taskListGrid-search-btn')
    taskListGridSearchBtn.click()
except NoSuchElementException:
    pass
time.sleep(5)
try:
    recordE = EC.presence_of_element_located((By.XPATH,
                                              "/html/body/main/div/div/div/div[3]/div/div[2]/div/div[1]/div/div["
                                              "3]/div[2]/div/div/div/div/div[1]/table/tbody/tr/td[4]/a"))
    WebDriverWait(driver, timeOut).until(recordE)
    record = driver.find_element(By.XPATH,
                                 "/html/body/main/div/div/div/div[3]/div/div[2]/div/div[1]/div/div[3]/div["
                                 "2]/div/div/div/div/div[1]/table/tbody/tr/td[4]/a")
    driver.execute_script("arguments[0].click();", record)
except TimeoutException:
    pass
time.sleep(5)
driver.switch_to.window(driver.window_handles[1])
driver.maximize_window()
time.sleep(5)
try:
    searchClaim = driver.find_element(By.XPATH,
                                      "/html/body/main/div/div/div/div[1]/div[2]/div[2]/div/div[2]/div/div["
                                      "1]/div/div[11]/div[1]/div[2]/button")
    searchClaim.click()
except NoSuchElementException:
    pass
time.sleep(5)
try:
    claimNumber = driver.find_element(By.XPATH,
                                      "/html/body/main/div/div/div/div[1]/div[2]/div[2]/div/div["
                                      "2]/div/div[1]/div/div[11]/div[2]/div/div/div/form/div/div["
                                      "1]/div/input")
    sheet = wb.active
    for cur_row in range(2, sheet.max_row + 1):
        claim_Number = sheet.cell(row=cur_row, column=1).value
        claimNumber.clear()
        claimNumber.send_keys(claim_Number)
        time.sleep(5)
        try:
            claimGridSearchBtn = driver.find_element(By.ID, 'claimGrid-search-btn')
            driver.execute_script("arguments[0].scrollIntoView();", claimGridSearchBtn)
            # claimGridSearchBtn.click()
        except NoSuchElementException:
            pass
        time.sleep(5)
        try:
            moneyBtn = driver.find_element(By.XPATH,
                                           "/html/body/main/div/div/div/div[1]/div[2]/div[2]/div/div[2]/div/div["
                                           "1]/div/div["
                                           "14]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td[20]/a")
            driver.implicitly_wait(10)
            ActionChains(driver).move_to_element(moneyBtn).click(moneyBtn).perform()
        except NoSuchElementException:
            pass
        time.sleep(5)
        try:
            sum_cost = sheet.cell(row=cur_row, column=2).value
            if sum_cost > 1300000:
                claimPayableAmount = driver.find_element(By.XPATH,
                                                         "/html/body/main/div/div/div/div[1]/div[2]/div[2]/div/div["
                                                         "2]/div/div[1]/div/div[17]/div/div/div[2]/div[1]/div["
                                                         "2]/div/div/div[4]/div[2]/div[1]/input")
                claimPayableAmount.clear()
                claimPayableAmount.send_keys("0")
                try:
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    addOther = driver.find_element(By.XPATH,
                                                   "/html/body/main/div/div/div/div[1]/div[2]/div[2]/div/div["
                                                   "2]/div/div["
                                                   "1]/div/div["
                                                   "17]/div/div/div[2]/div[2]/div[3]/div[2]/div/div/div/div/div[1]/a[2]")
                    addOther.click()
                except NoSuchElementException:
                    pass
                time.sleep(5)
                try:
                    selectItemContainer = Select(driver.find_element(By.ID,
                                                                     "itemId"))
                    selectItemContainer.select_by_visible_text("سایر")
                except NoSuchElementException:
                    pass
                time.sleep(5)

                requestAmount = sheet.cell(row=cur_row, column=3).value
                otherTotalAmountInputId = driver.find_element(By.ID,
                                                              "otherTotalAmountInputId")
                otherTotalAmountInputId.send_keys(requestAmount)
                time.sleep(5)
                try:
                    otherDescriptionInputId = driver.find_element(By.ID,
                                                                  "otherDescriptionInputId")
                    otherDescriptionInputId.send_keys("پرداخت از  محل صندوق")
                except NoSuchElementException:
                    pass
                time.sleep(5)
                try:
                    saveOtherBtn = driver.find_element(By.ID, 'btnSaveother')
                    saveOtherBtn.click()
                except NoSuchElementException:
                    pass
                time.sleep(5)
                try:
                    claimAssessmentBtn = driver.find_element(By.ID, 'btnClaimAssessment')
                    claimAssessmentBtn.click()
                except NoSuchElementException:
                    pass
                time.sleep(5)
                try:
                    closeBtn = driver.find_element(By.XPATH, "/html/body/main/div/div/div/div[1]/div["
                                                             "2]/div[2]/div/div[2]/div/div[1]/div/div["
                                                             "17]/div/div/div[1]/button")
                    closeBtn.click()
                except NoSuchElementException:
                    pass
                time.sleep(5)
            if sum_cost < 1300000:
                claimPayableAmount = driver.find_element(By.ID, 'claimPayableAmountInput')
                claimPayableAmount.clear()
                claim_Payable_Amount = 1300000 - sum_cost
                payable_Amount_str = str(claim_Payable_Amount)
                claimPayableAmount.send_keys(payable_Amount_str)
                try:
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    addOther = driver.find_element(By.XPATH,
                                                   "/html/body/main/div/div/div/div[1]/div[2]/div[2]/div/div["
                                                   "2]/div/div["
                                                   "1]/div/div["
                                                   "17]/div/div/div[2]/div[2]/div[3]/div[2]/div/div/div/div/div[1]/a[2]")
                    addOther.click()
                except NoSuchElementException:
                    pass
                time.sleep(5)
                try:
                    selectItemContainer = Select(driver.find_element(By.ID,
                                                                     "itemId"))
                    selectItemContainer.select_by_visible_text("سایر")
                except NoSuchElementException:
                    pass
                time.sleep(5)
                requestAmount = sheet.cell(row=cur_row, column=3).value
                other_Payable_Amount = requestAmount - claim_Payable_Amount
                otherTotalAmountInputId = driver.find_element(By.ID,
                                                              "otherTotalAmountInputId")
                otherTotalAmountInputId.clear()
                otherTotalAmount_str = str(other_Payable_Amount)
                otherTotalAmountInputId.send_keys(otherTotalAmount_str)
                time.sleep(5)
                try:
                    otherDescriptionInputId = driver.find_element(By.ID,
                                                                  "otherDescriptionInputId")
                    otherDescriptionInputId.send_keys("پرداخت از  محل صندوق")
                except NoSuchElementException:
                    pass
                time.sleep(5)
                try:
                    saveOtherBtn = driver.find_element(By.ID, 'btnSaveother')
                    saveOtherBtn.click()
                except NoSuchElementException:
                    pass
                time.sleep(5)
                try:
                    claimAssessmentBtn = driver.find_element(By.ID, 'btnClaimAssessment')
                    claimAssessmentBtn.click()
                except NoSuchElementException:
                    pass
                time.sleep(5)
        except NoSuchElementException:
            pass
except NoSuchElementException:
    pass
time.sleep(5)
time_now = datetime.now()
current_time = time_now.strftime("%H:%M:%S")
print('زمان پایان' + str(current_time))
